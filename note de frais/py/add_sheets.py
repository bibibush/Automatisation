import openpyxl

load_path = 'Modèle-de-note-de-frais.xlsx'
save_path = '27-28 Modèle-de-note-de-frais.xlsx'

wb = openpyxl.load_workbook(load_path)

months = ['Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
          'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin']
numero_months = [7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6]
this_year = 2027
next_year = 2028

for i in range(11 + 1):
    ws = wb.worksheets[i]
    
    if i <= 5:
        ws.title = months[i] + " " + str(this_year)
        ws['C2'] = months[i] + " " + str(this_year)
        for y in range(6, 36 + 1):
          ws[f'A{y}'].number_format = '[$-x-sysdate]dddd, mmmm dd, yyyy'
          ws[f'A{y}'].value = f'{this_year}-{numero_months[i]}-{y-5}'
          
          
    else:
        ws.title = months[i] + " " + str(next_year)
        ws['C2'] = months[i] + " " + str(next_year)
        for x in range(6, 36 + 1):
          ws[f'A{x}'].number_format = '[$-x-sysdate]dddd, mmmm dd, yyyy'
          ws[f'A{x}'].value = f'{next_year}-{numero_months[i]}-{x-5}'

wb.save(save_path)



    

        