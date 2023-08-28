import openpyxl

load_path = '07.Sept.23 MODELE Tableau CSE Spé Fetes.xlsx'
save_path = '07.Sept.23 MODELE Tableau CSE Spé Fetes.xlsx'

wb = openpyxl.load_workbook(load_path)
tarif_wb = openpyxl.load_workbook('tarif CSE.xlsx')
tarif_ws = tarif_wb.worksheets[1]

a = tarif_ws['B3'].value
b = tarif_ws['B4'].value
c = tarif_ws['B5'].value

d = tarif_ws['B7'].value
e = tarif_ws['B8'].value
f = tarif_ws['B9'].value

g = tarif_ws['B11'].value

h = tarif_ws['B13'].value
i = tarif_ws['B14'].value
j = tarif_ws['B15'].value

k = tarif_ws['B17'].value
l = tarif_ws['B18'].value
m = tarif_ws['B19'].value
n = tarif_ws['B20'].value
o = tarif_ws['B21'].value

p = tarif_ws['B23'].value

recap = wb.worksheets[1]
for z in range(6, 20 + 1):
    recap[f'T{z}'] = f'=C{z}*{a}+D{z}*{b}+E{z}*{c}+F{z}*{d}+G{z}*{e}+H{z}*{f}+I{z}*{g}+J{z}*{h}+K{z}*{i}+L{z}*{j}+M{z}*{k}+N{z}*{l}+O{z}*{m}+P{z}*{n}+Q{z}*{o}+R{z}*{p}'

for num in range(3, 17 + 1):
    ws = wb.worksheets[num - 1]
    for cn in range(6, 105 + 1):
        ws[f"V{cn}"] = f'=E{cn}*{a} + F{cn}*{b} + G{cn}*{c} + H{cn}*{d} + I{cn}*{e} + J{cn}*{f} + K{cn}*{g} + L{cn}*{h} + M{cn}*{i} + N{cn}*{j} + O{cn}*{k} + P{cn}*{l} + Q{cn}*{m} + R{cn}*{n} + S{cn}*{o} + T{cn}*{p}'

fac = wb['FAC']
fac['B4'] = a
fac['B5'] = b
fac['B6'] = c
fac['B7'] = d
fac['B8'] = e
fac['B9'] = f
fac['B10'] = g
fac['B11'] = h
fac['B12'] = i
fac['B13'] = j
fac['B14'] = k
fac['B15'] = l
fac['B16'] = m
fac['B17'] = n
fac['B18'] = o
fac['B20'] = p

fac['D4'] = round(a/1.055,2)
fac['D5'] = round(b/1.055,2)
fac['D6'] = round(c/1.055,2)
fac['D7'] = round(d/1.055,2)
fac['D8'] = round(e/1.055,2)
fac['D9'] = round(f/1.055,2)
fac['D10'] = round(g/1.055,2)
fac['D11'] = round(h/1.055,2)
fac['D12'] = round(i/1.055,2)
fac['D13'] = round(j/1.055,2)
fac['D14'] = round(k/1.055,2)
fac['D15'] = round(l/1.055,2)
fac['D16'] = round(m/1.055,2)
fac['D17'] = round(n/1.055,2)
fac['D18'] = round(o/1.055,2)
fac['D20'] = round(p/1.055,2)

wb.save(save_path)